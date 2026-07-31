# -*- coding: utf-8 -*-
"""
每日 Excel 练习题生成器
========================
每天由 DeepSeek 生成 5 道围绕「Excel 必学 16 函数」的练习题，
用 openpyxl 写成 xlsx（题目 / 答案 / 讲解 三 sheet），
同时输出 excel/daily.json 供主站渲染。

用法：
  LLM_API_KEY=xxx python gen_excel.py            # 真实调用 DeepSeek
  python gen_excel.py --mock                     # 离线测试（不联网）

输出：
  excel/<YYYY-MM-DD>.xlsx   当日文件（按北京时间命名）
  excel/latest.xlsx         最新文件（稳定拉取地址）
  excel/daily.json          结构化题目（date + questions[]），供前端渲染
"""
import os
import re
import sys
import json
import time
from datetime import datetime, timezone, timedelta

API_URL = "https://api.deepseek.com/chat/completions"
MODEL = "deepseek-chat"

# 与站点 state.skills.excel 对齐的 16 个必学函数
FUNC_LIST = ["IFERROR", "IF", "TODAY", "SUMIFS", "COUNTIFS", "SUMIF", "COUNTIF",
             "MAX", "ROUND", "XLOOKUP", "VLOOKUP", "DATEDIF", "RIGHT", "MID",
             "INDEX", "MATCH"]

CJK = re.compile(r"[\u4e00-\u9fff]")


def log(*a):
    print("[gen_excel]", *a, flush=True)


# ---------- LLM（与 gen_daily.py 同款，无需第三方库） ----------
def llm(messages, max_tokens=2200, temperature=0.9):
    import urllib.request
    api_key = os.environ.get("LLM_API_KEY")
    if not api_key:
        raise SystemExit("LLM_API_KEY 环境变量缺失")
    data = json.dumps({
        "model": MODEL,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }).encode("utf-8")
    req = urllib.request.Request(API_URL, data=data, headers={
        "Authorization": "Bearer " + api_key,
        "Content-Type": "application/json",
    })
    last_err = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=90) as r:
                resp = json.loads(r.read().decode("utf-8"))
            return resp["choices"][0]["message"]["content"]
        except Exception as e:  # noqa
            last_err = e
            log("LLM 请求失败(重试 %d): %s" % (attempt + 1, e))
            time.sleep(3)
    raise SystemExit("LLM 调用最终失败: %s" % last_err)


def extract_json(text):
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    start = None
    for i, ch in enumerate(text):
        if ch in "[{":
            start = i
            break
    if start is None:
        return None
    opener, closer = text[start], ("}" if text[start] == "{" else "]")
    depth = 0
    for i in range(start, len(text)):
        if text[i] == opener:
            depth += 1
        elif text[i] == closer:
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return None


def gen_json(prompt, max_tokens=2400):
    for _ in range(3):
        raw = llm([{"role": "system",
                    "content": "你是严谨的 Excel 教学出题人，只输出符合要求的 JSON，不要任何解释。"},
                   {"role": "user", "content": prompt}], max_tokens=max_tokens)
        js = extract_json(raw)
        if js:
            try:
                return json.loads(js)
            except Exception as e:  # noqa
                log("JSON 解析失败，重试:", e)
    raise SystemExit("无法从模型回复解析出 JSON")


# ---------- 题目生成 ----------
def make_questions(n=5):
    func_str = "、".join(FUNC_LIST)
    prompt = (
        "请出 %d 道 Microsoft Excel 实操练习题，面向「零基础学 Excel 的成人」，"
        "围绕以下 16 个必学函数出题：%s。\n"
        "要求：\n"
        "1. 每道题练习 1~2 个函数，5 道题尽量覆盖不同函数、难度从易到难；\n"
        "2. 题目要贴近真实生活/办公/物流场景（如台账、重量汇总、到港天数、查表匹配等），不要抽象；\n"
        "3. 每道题必须给出一份可直接粘贴进 Excel 的「样例数据」：一个二维数组 data，"
        "第一行是表头（列名），其后是若干数据行，所有内容都是字符串或数字，不要公式；\n"
        "4. 给出明确「任务」描述（让用户在某单元格输入什么公式/得到什么结果）；\n"
        "5. 给出正确「答案」：若是公式请写以 = 开头的 Excel 公式，若是结果写具体值；\n"
        "6. 给出「讲解」：用中文说明用到的函数语法和思路。\n\n"
        "严格按以下 JSON 数组返回，字段含义见注释（不要返回注释）：\n"
        "[\n"
        '  {"q":"第1题标题","func":"SUMIFS","desc":"题目背景与要求描述","data":[["列A","列B"],["值1","值2"]],'
        '"task":"请在哪个单元格输入什么","answer":"=SUMIFS(...)","explain":"讲解..."}\n'
        "]\n"
        "只返回 JSON 数组，不要其他内容。" % (n, func_str)
    )
    items = gen_json(prompt)
    good, dropped = [], 0
    for it in items:
        try:
            q = str(it.get("q", "")).strip()
            func = str(it.get("func", "")).strip().upper()
            desc = str(it.get("desc", "")).strip()
            data = it.get("data", [])
            task = str(it.get("task", "")).strip()
            answer = str(it.get("answer", "")).strip()
            explain = str(it.get("explain", "")).strip()
            if not (q and desc and task and answer and explain):
                dropped += 1; continue
            if not isinstance(data, list) or len(data) < 2:
                dropped += 1; continue
            # 校验 data 是二维、首行为表头
            if not all(isinstance(r, list) for r in data):
                dropped += 1; continue
            if func and func not in FUNC_LIST:
                func = ""  # 不强制，留空也可
            good.append({"q": q, "func": func, "desc": desc,
                         "data": data, "task": task, "answer": answer, "explain": explain})
        except Exception:
            dropped += 1
    log("questions 生成 %d，保留 %d，丢弃 %d" % (len(items), len(good), dropped))
    return good


# ---------- 写出 xlsx ----------
def write_xlsx(questions, date_str, out_dir):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws_q = wb.active
    ws_q.title = "题目"
    ws_a = wb.create_sheet("答案")
    ws_e = wb.create_sheet("讲解")

    head_font = Font(bold=True, size=12, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F6FED")
    title_font = Font(bold=True, size=13, color="2F6FED")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = wrap
            cell.border = border

    r = 1
    for i, q in enumerate(questions, 1):
        # ---- 题目 sheet ----
        ws_q.cell(row=r, column=1, value="第 %d 题：%s" % (i, q["q"])).font = title_font
        r += 1
        if q["func"]:
            ws_q.cell(row=r, column=1, value="练习函数：" + q["func"])
            r += 1
        ws_q.cell(row=r, column=1, value="要求：" + q["desc"]).alignment = wrap
        r += 1
        ws_q.cell(row=r, column=1, value="任务：" + q["task"]).alignment = wrap
        r += 1
        # 样例数据表（首行表头）
        data = q["data"]
        ncols = max(len(row) for row in data)
        for ri, row in enumerate(data):
            for ci, val in enumerate(row, 1):
                cell = ws_q.cell(row=r + ri, column=ci, value=val)
                cell.border = border
                cell.alignment = wrap
            if ri == 0:
                style_header_row(ws_q, r, ncols)
        r += len(data)
        # 预留作答区
        ws_q.cell(row=r, column=1, value="↓ 你的公式 / 答案（在下方单元格输入）：").font = Font(italic=True, color="888888")
        r += 2
        # ---- 答案 sheet ----
        ws_a.cell(row=(i * 4 - 3), column=1, value="第 %d 题：%s" % (i, q["q"])).font = title_font
        ws_a.cell(row=(i * 4 - 2), column=1, value="答案：").font = Font(bold=True)
        ws_a.cell(row=(i * 4 - 2), column=2, value=q["answer"]).alignment = wrap
        ws_a.cell(row=(i * 4 - 1), column=1, value="（可直接复制公式到题目文件对应单元格验证）")
        # ---- 讲解 sheet ----
        ws_e.cell(row=(i * 3 - 2), column=1, value="第 %d 题：%s" % (i, q["q"])).font = title_font
        ws_e.cell(row=(i * 3 - 1), column=1, value="讲解：").font = Font(bold=True)
        ws_e.cell(row=(i * 3 - 1), column=2, value=q["explain"]).alignment = wrap

    ws_q.column_dimensions["A"].width = 26
    for col in "BCDE":
        ws_q.column_dimensions[col].width = 18
    ws_a.column_dimensions["A"].width = 16
    ws_a.column_dimensions["B"].width = 80
    ws_e.column_dimensions["A"].width = 16
    ws_e.column_dimensions["B"].width = 90

    os.makedirs(out_dir, exist_ok=True)
    path_dated = os.path.join(out_dir, date_str + ".xlsx")
    path_latest = os.path.join(out_dir, "latest.xlsx")
    wb.save(path_dated)
    wb.save(path_latest)
    log("xlsx 已写出：", path_dated, "（最新副本 latest.xlsx）")


# ---------- 写出 daily.json ----------
def write_json(questions, date_str, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    payload = {"date": date_str, "questions": questions}
    path = os.path.join(out_dir, "daily.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    log("daily.json 已写出，共", len(questions), "题，日期", date_str)


# ---------- 本地时区（北京时间） ----------
def beijing_date():
    return datetime.now(timezone(timedelta(hours=8))).strftime("%Y-%m-%d")


def mock_questions():
    return [
        {"q": "汇总某航线空运总重量", "func": "SUMIFS",
         "desc": "下表是若干票货物的运输方式与重量(kg)，请汇总「空运」方式的总重量。",
         "data": [["序号", "运输方式", "重量(kg)"], ["1", "空运", "320"], ["2", "海运", "1500"],
                  ["3", "空运", "480"], ["4", "空运", "260"], ["5", "海运", "900"]],
         "task": "在 E2 单元格输入公式，汇总运输方式=空运的重量合计。",
         "answer": "=SUMIFS(C2:C6, B2:B6, \"空运\")",
         "explain": "SUMIFS(求和区, 条件区1, 条件1)：对满足「运输方式=空运」的行，把对应重量相加，结果为 320+480+260=1060。"},
        {"q": "判断是否已超期", "func": "IF + TODAY",
         "desc": "下表是各票货物预计到港日，请判断是否已超期（到港日 < 今天 为超期）。",
         "data": [["出货编号", "预计到港日"], ["BKG001", "2026-08-05"], ["BKG002", "2026-07-28"],
                  ["BKG003", "2026-08-10"]],
         "task": "在 C2 输入公式，超期显示「已超期」，否则显示「正常」。",
         "answer": "=IF(B2<TODAY(), \"已超期\", \"正常\")",
         "explain": "IF(条件, 成立值, 不成立值)：TODAY() 返回今天日期，日期可直接比较大小。"},
        {"q": "按柜号查目的港", "func": "VLOOKUP",
         "desc": "左侧为出货台账（编号/目的港），右侧给出若干柜号，请查出对应目的港。",
         "data": [["出货编号", "目的港"], ["BKG001", "LAX"], ["BKG002", "NYC"], ["BKG003", "HKG"]],
         "task": "在 D2 输入公式，按 A 列编号查目的港（精确匹配）。",
         "answer": "=VLOOKUP(C2, A:B, 2, FALSE)",
         "explain": "VLOOKUP(查找值, 表格区, 返回列号, 0)：在第1列找 C2，返回第2列的值；FALSE 表示精确匹配。"},
        {"q": "提取出货编号年份", "func": "MID",
         "desc": "出货编号形如 BKG2026003，请提取其中的年份(2026)。",
         "data": [["出货编号", "年份"], ["BKG2026001", ""], ["BKG2026002", ""]],
         "task": "在 B2 输入公式提取年份（第4位起，长度4）。",
         "answer": "=MID(A2, 4, 4)",
         "explain": "MID(文本, 起始位置, 长度)：从第4个字符开始取4位，得到 2026。"},
        {"q": "计算运输时长(天)", "func": "DATEDIF",
         "desc": "已知开船日与到港日，请计算运输天数。",
         "data": [["开船日", "到港日", "天数"], ["2026-07-20", "2026-07-27", ""],
                  ["2026-08-01", "2026-08-12", ""]],
         "task": "在 C2 输入公式计算两个日期相差的天数。",
         "answer": "=DATEDIF(A2, B2, \"D\")",
         "explain": "DATEDIF(开始, 结束, \"D\")：返回两个日期之间相差的天数。"},
    ]


def main():
    date_str = beijing_date()
    out_dir = "excel"
    use_mock = "--mock" in sys.argv
    if use_mock:
        log("【离线模式】使用 mock 题目，不调用 LLM")
        questions = mock_questions()
    else:
        questions = make_questions(5)
    if not questions:
        raise SystemExit("未生成任何有效题目")
    write_xlsx(questions, date_str, out_dir)
    write_json(questions, date_str, out_dir)
    log("完成。日期=%s，题目数=%d" % (date_str, len(questions)))


if __name__ == "__main__":
    main()
